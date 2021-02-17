from datetime import date
import datetime
import pandas as pd
import praw
from praw.models import MoreComments
import requests
import csv
import openpyxl
from openpyxl import load_workbook
from collections import defaultdict, Counter


#authentication tokens
reddit = praw.Reddit(client_id='Hoh5COYowZyKEw', client_secret='Xt5uF0VRX9HyjopbovAlwWxe2CcPMg', user_agent='WSB_Stock_Mentions')

#globalDictionary to update stock count
stock_dictionary= {"stock_name": "count"}
column_names_limited_stocks_csv=['NASDAQ']

def percent(x,y):
    percent = ( ( (x-y) / y ) *100)
    return round(percent,1)

#updates count in csv
#TODO - find how to change the count in each column (should just be a +1 for each)
#     - reset day, week, month, and year on correct day
def reset_day():
    curr_hour = pd.Timestamp.now().hour
    if curr_hour == 0:
        return True
    else:
        return False

def reset_week():
    #today = today.date()
    if ((datetime.datetime.today().weekday() == "Sunday") and reset_day()):
        return True
    else:
        return False

def reset_open_close():
    curr_hour = pd.Timestamp.now().hour
    if(curr_hour==9 and reset_week()):
        return True
    else:
        return False



def find_trending():
    workbook = load_workbook('listed_stocks.xlsx')

    hourly_worksheet = workbook["hourly"]
    daily_worksheet = workbook["daily"]
    weekly_worksheet = workbook["weekly"]
    open_closed_worksheet = workbook["open_closed"]

    hourly_column_curr = hourly_worksheet['C']
    daily_column_curr = daily_worksheet['C']
    weekly_column_curr = weekly_worksheet['C']
    open_closed_column_curr = open_closed_worksheet['C']

    hourly_column_prev = hourly_worksheet['D']
    daily_column_prev = daily_worksheet['D']
    weekly_column_prev = weekly_worksheet['D']
    open_closed_column_prev = open_closed_worksheet['D']

    hourly_list_curr = [hourly_column_curr[x].value for x in range(1,len(hourly_column_curr))]
    daily_list_curr = [daily_column_curr[x].value for x in range(1,len(daily_column_curr))]
    weekly_list_curr = [weekly_column_curr[x].value for x in range(1,len(weekly_column_curr))]
    open_closed_list_curr = [open_closed_column_curr[x].value for x in range(1,len(open_closed_column_curr))]

    hourly_list_prev = [hourly_column_prev[x].value for x in range(1,len(hourly_column_prev))]
    daily_list_prev = [daily_column_prev[x].value for x in range(1,len(daily_column_prev))]
    weekly_list_prev = [weekly_column_prev[x].value for x in range(1,len(weekly_column_prev))]
    open_closed_list_prev = [open_closed_column_prev[x].value for x in range(1,len(open_closed_column_prev))]

    NASDAQ_list = get_stock_list("NASDAQ")
    NASDAQ_list.pop(0)

    hourly_trending_values = []
    daily_trending_values = []
    weekly_trending_values = []
    open_closed_trending_values = []

    for i in range(len(NASDAQ_list)):
        if (hourly_list_prev[i] > 9):

            hourly_trending_values.append(percent(hourly_list_curr[i],hourly_list_prev[i]))
        else:
            hourly_trending_values.append(0)

        if (daily_list_prev[i] > 9):
            daily_trending_values.append(percent(daily_list_curr[i],daily_list_prev[i]))
        else:
            daily_trending_values.append(0)

        if (weekly_list_prev[i] > 9):
            weekly_trending_values.append(percent(weekly_list_curr[i],weekly_list_prev[i]))
        else:
            weekly_trending_values.append(0)

        if (open_closed_list_prev[i] > 9):
            open_closed_trending_values.append(percent(open_closed_list_curr[i],open_closed_list_prev[i]))
        else:
            open_closed_trending_values.append(0)


    hourly_values_dict = dict(zip(NASDAQ_list, hourly_trending_values))
    daily_values_dict = dict(zip(NASDAQ_list, daily_trending_values))
    weekly_values_dict = dict(zip(NASDAQ_list, weekly_trending_values))
    open_closed_values_dict = dict(zip(NASDAQ_list, open_closed_trending_values))

    hourly_values_dict =  sorted( hourly_values_dict.items(), key=lambda x: x[1], reverse=True)
    daily_values_dict = sorted( daily_values_dict.items(), key=lambda x: x[1], reverse=True)
    weekly_values_dict =  sorted( weekly_values_dict.items(), key=lambda x: x[1], reverse=True)
    open_closed_values_dict =  sorted( open_closed_values_dict.items(), key=lambda x: x[1], reverse=True)


    for i in range(9):
        row_value = i + 2
        cell_reference_hourly_top_stock = hourly_worksheet.cell(row=row_value,column=9)
        cell_reference_hourly_top_stock.value = hourly_values_dict[i][0]
        cell_reference_hourly_top_count = hourly_worksheet.cell(row=row_value,column=10)
        cell_reference_hourly_top_count.value = hourly_values_dict[i][1]

        if(reset_day()):
            cell_reference_daily_top_stock = daily_worksheet.cell(row=row_value,column=9)
            cell_reference_daily_top_stock.value = daily_values_dict[i][0]
            cell_reference_daily_top_count = daily_worksheet.cell(row=row_value,column=10)
            cell_reference_daily_top_count.value = daily_values_dict[i][1]
        if(reset_week()):
            cell_reference_weekly_top_stock = weekly_worksheet.cell(row=row_value,column=9)
            cell_reference_weekly_top_stock.value = weekly_values_dict[i][0]
            cell_reference_weekly_top_count = weekly_worksheet.cell(row=row_value,column=10)
            cell_reference_weekly_top_count.value = weekly_values_dict[i][1]
        if(reset_open_close()):
            cell_reference_open_closed_top_stock = open_closed_worksheet.cell(row=row_value,column=9)
            cell_reference_open_closed_top_stock.value = open_closed_values_dict[i][0]
            cell_reference_open_closed_top_count = open_closed_worksheet.cell(row=row_value,column=10)
            cell_reference_open_closed_top_count.value = open_closed_values_dict[i][1]

    workbook.save('listed_stocks.xlsx')
    workbook.close()


#Column C (3): Curr,
#Column F (6); top mentioned

def find_top_mentioned():
    workbook = load_workbook('listed_stocks.xlsx')

    hourly_worksheet = workbook["hourly"]
    daily_worksheet = workbook["daily"]
    weekly_worksheet = workbook["weekly"]
    open_closed_worksheet = workbook["open_closed"]

    hourly_column_curr = hourly_worksheet['C']
    daily_column_curr = daily_worksheet['C']
    weekly_column_curr = weekly_worksheet['C']
    open_closed_column_curr = open_closed_worksheet['C']

    NASDAQ_list = get_stock_list("NASDAQ")


    hourly_list = [hourly_column_curr[x].value for x in range(1,len(hourly_column_curr))]
    daily_list = [daily_column_curr[x].value for x in range(1,len(daily_column_curr))]
    weekly_list = [weekly_column_curr[x].value for x in range(1,len(weekly_column_curr))]
    open_closed_list = [open_closed_column_curr[x].value for x in range(1,len(open_closed_column_curr))]


    #necessary to pop the column headers off so we can sort, if we don't we try to compare the string "top mentioned" to ints
    NASDAQ_list.pop(0)
    hourly_values_dict = dict(zip(NASDAQ_list, hourly_list))
    daily_values_dict = dict(zip(NASDAQ_list, daily_list))
    weekly_values_dict = dict(zip(NASDAQ_list, weekly_list))
    open_closed_values_dict = dict(zip(NASDAQ_list, open_closed_list))

    hourly_values_dict =  sorted( hourly_values_dict.items(), key=lambda x: x[1], reverse=True)
    daily_values_dict = sorted( daily_values_dict.items(), key=lambda x: x[1], reverse=True)
    weekly_values_dict =  sorted( weekly_values_dict.items(), key=lambda x: x[1], reverse=True)
    open_closed_values_dict =  sorted( open_closed_values_dict.items(), key=lambda x: x[1], reverse=True)


    for i in range(9):
        row_value = i + 2
        cell_reference_hourly_top_stock = hourly_worksheet.cell(row=row_value,column=6)
        cell_reference_hourly_top_stock.value = hourly_values_dict[i][0]
        cell_reference_hourly_top_count = hourly_worksheet.cell(row=row_value,column=7)
        cell_reference_hourly_top_count.value = hourly_values_dict[i][1]


        cell_reference_daily_top_stock = daily_worksheet.cell(row=row_value,column=6)
        cell_reference_daily_top_stock.value = daily_values_dict[i][0]
        cell_reference_daily_top_count = daily_worksheet.cell(row=row_value,column=7)
        cell_reference_daily_top_count.value = daily_values_dict[i][1]

        cell_reference_weekly_top_stock = weekly_worksheet.cell(row=row_value,column=6)
        cell_reference_weekly_top_stock.value = weekly_values_dict[i][0]
        cell_reference_weekly_top_count = weekly_worksheet.cell(row=row_value,column=7)
        cell_reference_weekly_top_count.value = weekly_values_dict[i][1]

        cell_reference_open_closed_top_stock = open_closed_worksheet.cell(row=row_value,column=6)
        cell_reference_open_closed_top_stock.value = open_closed_values_dict[i][0]
        cell_reference_open_closed_top_count = open_closed_worksheet.cell(row=row_value,column=7)
        cell_reference_open_closed_top_count.value = open_closed_values_dict[i][1]


    workbook.save('listed_stocks.xlsx')
    workbook.close()



def check_resets():

    NASDAQ_list = get_stock_list('NASDAQ')
    print("check resets entered")
    listed_stocks_workbook = load_workbook('listed_stocks.xlsx')
    open_close_worksheet = listed_stocks_workbook['open_closed']
    hourly_worksheet = listed_stocks_workbook['hourly']
    daily_worksheet = listed_stocks_workbook['daily']
    weekly_worksheet = listed_stocks_workbook['weekly']
#Curr = Column C (or 3)https://www.youtube.com/watch?v=ft4Wiy4LcCg&feature=emb_logo
#Prev = Column D
    #at midnight we will reset current count and move curr to prev
    if(reset_day()):
        for i in range(2, len(NASDAQ_list)+1):
            print("reset_day")
            cell_reference_daily_curr = daily_worksheet.cell(row=i,column=3)
            cell_reference_daily_prev = daily_worksheet.cell(row=i, column=4)
            cell_reference_daily_prev.value = cell_reference_daily_curr.value
            cell_reference_daily_curr.value = 0


    #mon-fri we will reset when the market opens.
    if(reset_open_close()):

        for i in range(2, len(NASDAQ_list)+1):
            print("reset Open_Close")
            cell_reference_open_close_curr = open_close_worksheet.cell(row=i,column=3)
            cell_reference_open_close_prev = open_close_worksheet.cell(row=i, column=4)
            cell_reference_open_close_prev.value = cell_reference_open_close_curr.value
            cell_reference_open_close_curr.value = 0

#reset Week on monday
    if(reset_week()):
        print("reset _ week")
        for i in range(2, len(NASDAQ_list)+1):
            cell_reference_week_curr = weekly_worksheet.cell(row=i,column=3)
            cell_reference_week_prev = weekly_worksheet.cell(row=i, column=4)
            cell_reference_week_prev.value= cell_reference_week_curr.value
            cell_reference_week_curr.value = 0

#this script runs hourly so we will always reset hour Counter
    for i in range(2, len(NASDAQ_list)+1):

        cell_reference_hour_curr = hourly_worksheet.cell(row=i,column=3)
        cell_reference_hour_prev = hourly_worksheet.cell(row=i, column=4)
        if(cell_reference_hour_curr.value != 0 or cell_reference_hour_prev.value!=0):
            cell_reference_hour_prev.value = cell_reference_hour_curr.value
            cell_reference_hour_curr.value = 0

    listed_stocks_workbook.save('listed_stocks.xlsx')
    listed_stocks_workbook.close()





def update_stock_counts():
    check_resets()

    NASDAQ_list = get_stock_list('NASDAQ')
    listed_stocks_workbook = load_workbook('listed_stocks.xlsx')
    open_close_worksheet = listed_stocks_workbook['open_closed']
    hourly_worksheet = listed_stocks_workbook['hourly']
    daily_worksheet = listed_stocks_workbook['daily']
    weekly_worksheet = listed_stocks_workbook['weekly']
    #Curr = Column C

    for stock in stock_dictionary:
        if stock in NASDAQ_list:

            stock_row = NASDAQ_list.index(stock) + 1
            cell_reference_daily = daily_worksheet.cell(row=stock_row,column=3)
            cell_reference_weekly = weekly_worksheet.cell(row=stock_row,column=3)
            cell_reference_open_close = open_close_worksheet.cell(row=stock_row,column=3)
            cell_reference_hourly = hourly_worksheet.cell(row=stock_row,column=3)

            cell_reference_daily.value = cell_reference_daily.value + stock_dictionary[stock]
            cell_reference_weekly.value = cell_reference_weekly.value + stock_dictionary.get(stock)
            cell_reference_open_close.value = cell_reference_open_close.value + stock_dictionary.get(stock)
            cell_reference_hourly.value = cell_reference_hourly.value + stock_dictionary.get(stock)
    listed_stocks_workbook.save('listed_stocks.xlsx')
    listed_stocks_workbook.close()
    #update metric for open_close, hourly, daily, and weekly
    #wipe open)close at each open (lets just do 9-4)
    #will need a time and date class


def get_stock_list(stock_type):

    wb = load_workbook('listed_stocks.xlsx')
    ws =wb["listed_stocks"]
    column = ws['A']
    column_list = [column[x].value for x in range(len(column))]

    return column_list
#    if(stock_type=='NASDAQ'):
#        #print  df['NASDAQ'].to_list()
##
#    elif(stock_type=='CQS'):
#        return df['CQS'].to_list()
#    else:
#        return df['ACT'].to_list()

#removes nonEmoji characters
def ignore_emoji(content):
    return content.encode('ascii','ignore').decode('ascii')

#scans actual text passage
#creates NASDAQ_list once to be used in scan text_list
NASDAQ_list=get_stock_list('NASDAQ')

def scan_text_for_stocks(text):

    text = ignore_emoji(text)

    #CQS_list = get_stock_list('CQS')
    #ACT_list = get_stock_list('ACT')


    text.replace('$','')
    text_list = text.split()

    for word in text_list:
        if word in (NASDAQ_list): #or CQS_list or ACT_list):
            if word in stock_dictionary:
                stock_dictionary[word] = stock_dictionary.get(word)+1 #.update(word = stock_dictionary.get(word)+1)

            else:
                stock_dictionary[word]=1


#TODO SPLIT STRING INTO ARRAY OF WORDS AND ADD TO STOCK DICTIONARY, IF EXISTS THEN ++
#pulls data from postArchive.txt or commentArchive.txt to create a list that filter methods can use to compare.
def generate_archive_list(input):

    if(input=='post'):
        archive_file = open("postArchive.txt",'r')
    else:
        archive_file = open('commentArchive.txt','r')
    archive_list= [(line.strip()) for line in archive_file]
    archive_file.close()
    return archive_list


#parses through all comments, filters ones that have NOT been scanned and passes those to scan_text_for_stocks
def filter_comments(comment_forest):

    comment_archive_file = open('commentArchive.txt','a')
    comment_archive_list = generate_archive_list('comment')

    for top_level_comment in comment_forest:
        if(top_level_comment.id not in comment_archive_list):
            if isinstance(top_level_comment, MoreComments):
                continue
                #print(ignore_emoji(top_level_comment.body))
            scan_text_for_stocks(top_level_comment.body)
            comment_archive_file.write('\n' + top_level_comment.id)



#parses through all posts, filters ones that have NOT been scanned and passes those to scan_text_for_stocks
#passes each posts comments to filter_comments to be checked separately
def filter_posts(wsb_posts):

    post_archive_file = open('postArchive.txt','a')
    post_archive_list = generate_archive_list('post')

    for post_id in wsb_posts:
        reddit_post = reddit.submission(id=post_id)
        if(post_id not in post_archive_list):
            post_archive_file.write("\n" + post_id)
            reddit_post_title = reddit_post.title
            reddit_post_selftext = reddit_post.selftext
            scan_text_for_stocks(reddit_post_title)
            scan_text_for_stocks(reddit_post_selftext)
        filter_comments(reddit_post.comments)
    post_archive_file.close()





#grabs top 50 posts IDs in each category and passes them through filter posts
def create_post_list():


    wsb_posts = []
    wsb_subreddit = reddit.subreddit('wallstreetbets')
    #successfully grabs top 100 posts
    for post in wsb_subreddit.hot(limit=25):
        wsb_posts.append(post.id)
    for post in wsb_subreddit.top('day', limit=25):
        wsb_posts.append(post.id)
    for post in wsb_subreddit.new(limit=25):
        wsb_posts.append(post.id)
    for post in wsb_subreddit.rising(limit=25):
        wsb_posts.append(post.id)


    return wsb_posts

def main():
    print("scraping wsb....")
    filter_posts(create_post_list())
    print(stock_dictionary)
    update_stock_counts()
    find_top_mentioned()
    find_trending()




main()
